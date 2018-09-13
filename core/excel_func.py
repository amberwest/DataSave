# -*- coding:utf-8 -*-
# Time: 2018/6/14 09:49
import time
import os
from openpyxl import load_workbook
import pandas as pd

from DataSave import debug_logger, error_logger, path



class ExcelFunc(object):
    def __init__(self, file='', sheet_name='sheet1'):
        if file:
            self.file = path + '{}.xlsx'.format(file)
        else:
            self.file = path + '{}.xlsx'.format(time.strftime('%Y%m%d'))
        self.file_is_exist = os.path.exists(file)
        self.sheet_name = sheet_name

    def save(self, data):
        if not isinstance(data, dict):
            error_logger.logger.error("data必须是字典")

        columns = sorted(data.keys())
        df = pd.DataFrame([data], columns=columns)

        if not self.file_is_exist:
            writer = pd.ExcelWriter(self.file, engine='openpyxl')
            df.to_excel(writer, sheet_name=self.sheet_name, index=False, columns=columns)
            writer.save()
            debug_logger.logger.debug('save data to excel %s' % data)

        else:
            book = load_workbook(self.file)
            row_start = book[self.sheet_name].max_row
            writer = pd.ExcelWriter(self.file, engine='openpyxl')
            writer.book= book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(
                writer,
                sheet_name=self.sheet_name,
                index=False,
                header=False,
                startrow=row_start,
                columns=columns,
            )
            writer.save()
            debug_logger.logger.debug('save data to excel %s' % data)
